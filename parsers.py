"""
parsers.py — Reads Ovvi Fleet QC spreadsheet and outputs normalized error data.

THIS IS THE ONLY FILE THAT NEEDS TO CHANGE WHEN THE INPUT FORMAT CHANGES.
Everything else (app.py) works off the normalized DataFrame this produces.

Output DataFrame columns:
    date            — datetime.date
    error_code      — str, e.g. "A5-002" or "E71-65547"
    count           — int, number of occurrences that day
    serial_number   — str, 12-char hex MAC address
    unit_name       — str, human-readable name
    firmware_version— str, e.g. "ovvi-fw-v1.0.7"
    unit_type       — str, e.g. "Customer", "F&F Tester", "Influencer"
    error_category  — str, "A-Code (Alert)", "E-Code (Firmware)", or "M-Code (Message)"
"""

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path


# ---------------------------------------------------------------------------
# Error code definitions — maps code to human-readable short name
# Parsed from the "Error Code Definition" sheet, plus E-code lookup
# ---------------------------------------------------------------------------

# These are the A-codes from the Error Code Definition sheet.
# E-codes are more granular (state + error number) and decoded differently.
A_CODE_NAMES = {
    "A7-001": "Load Cell Error",
    "A1-001": "Can Dispense Failure",
    "A5-001": "Can Dispose Failure",
    "A5-002": "Empty-Chute Full/Jammed",
    "A7-002": "Carousel Jam",
    "A7-003": "Tilt Error",
    "A4-001": "Lid Retrieval Failed",
    "A7-004": "Power Lost During Dispense/Dispose",
    "A1-002": "New-Can Chute Stack Dropped",
    "A5-003": "Empty-Can Chute Stack Dropped",
    "A6-001": "Chute Not Present",
    "A2-001": "Open Can Failure / Can Rejected",
    "A7-005": "Blade Replacement Needed",
    "A7-006": "Chutes Lifted But Issue Detected",
    "A1-003": "Unopened Can in Empty Chute",
    "A8-002": "Feeding Weighing Error",
    "A6-002": "Chute Process Incomplete",
}

M_CODE_NAMES = {
    "M1-002": "1 Can Left",
    "M1-003": "No More Cans",
    "M3-002": "Cat Not Eating",
    "M5-001": "Chute Almost Full",
    "M5-002": "Chute Full",
    "M7-001": "Connection Error",
    "M5-003": "Cans Detected in Chute",
}

# Codes that should be ignored per the spreadsheet notes
IGNORE_CODES = {"M1-002", "M1-003", "M3-002", "M5-001", "M5-002", "A6-001"}


def get_error_name(code: str) -> str:
    """Return human-readable name for an error code."""
    code = code.strip()
    if code in A_CODE_NAMES:
        return A_CODE_NAMES[code]
    if code in M_CODE_NAMES:
        return M_CODE_NAMES[code]
    if code.startswith("E"):
        return f"Firmware Error {code}"
    return code


def get_error_category(code: str) -> str:
    """Classify error code into broad category."""
    code = code.strip()
    if code.startswith("A"):
        return "A-Code (Alert)"
    elif code.startswith("E"):
        return "E-Code (Firmware)"
    elif code.startswith("M"):
        return "M-Code (Message)"
    return "Unknown"


# ---------------------------------------------------------------------------
# Horizontal sheet parser
# ---------------------------------------------------------------------------

def _parse_horizontal_sheet(ws, sheet_name: str) -> list[dict]:
    """
    Parse one of the horizontal error trend sheets.
    
    Each unit gets 1+ rows. A row contains:
      - Metadata in early columns (name, type, SN, firmware version)
      - Repeating (date, error_code, count) triplets going right
    
    Continuation rows for the same unit may omit some metadata fields.
    """
    events = []
    current_name = None
    current_sn = None
    current_fw = None
    current_type = None

    # Known non-name string values that appear in early columns
    SKIP_STRINGS = {
        "Online", "Offline", "Feeding", "Not Feeding", "Notes", "Return",
        "Type", "Grams", "Reason", "Error Type", "Individual", "Name",
        "Serial Numbers:2", "Firmware Version", "Date", "Error Code",
        "Total Per Code", "Error Code Trend", "Error Code Trend for Customers",
    }
    TYPE_STRINGS = {"Customer", "F&F Tester", "Indiegogo", "Influencer", "Contractor", "Factory"}

    for row in ws.iter_rows(min_row=8, values_only=True):
        vals = list(row)
        if not any(v is not None for v in vals):
            continue

        # --- Extract metadata from early columns ---
        row_name = None
        row_sn = None
        row_fw = None
        row_type = None

        for v in vals[:12]:
            if v is None:
                continue
            if isinstance(v, (int, float)) and not isinstance(v, datetime):
                continue  # skip row numbers, gram counts, etc.
            if isinstance(v, str):
                vs = v.strip()
                if not vs:
                    continue
                if vs in SKIP_STRINGS:
                    continue
                if vs in TYPE_STRINGS:
                    row_type = vs
                elif vs.startswith("ovvi-fw"):
                    row_fw = vs
                elif len(vs) == 12 and all(c in "0123456789ABCDEFabcdef" for c in vs):
                    row_sn = vs.upper()
                elif (
                    len(vs) > 12
                    and vs.replace(" ", "").isalnum()
                    and all(c in "0123456789ABCDEFabcdef " for c in vs)
                    and len(vs.replace(" ", "")) == 12
                ):
                    row_sn = vs.replace(" ", "").upper()
                elif len(vs) > 2 and row_name is None:
                    row_name = vs

        # Update running state (continuation rows inherit from previous)
        if row_name:
            current_name = row_name
        if row_sn:
            current_sn = row_sn
        if row_fw:
            current_fw = row_fw
        if row_type:
            current_type = row_type

        # --- Extract date/code/count triplets ---
        i = 0
        while i < len(vals) - 2:
            if (
                isinstance(vals[i], datetime)
                and isinstance(vals[i + 1], str)
                and vals[i + 1].strip()
                and vals[i + 2] is not None
            ):
                try:
                    code = vals[i + 1].strip()
                    count = int(vals[i + 2])
                    if count > 0 and len(code) >= 2:
                        events.append({
                            "date": vals[i].date(),
                            "error_code": code,
                            "count": count,
                            "serial_number": current_sn,
                            "unit_name": current_name,
                            "firmware_version": current_fw,
                            "unit_type": current_type,
                            "source_sheet": sheet_name,
                        })
                except (ValueError, TypeError):
                    pass
            i += 1

    return events


# ---------------------------------------------------------------------------
# Error code definitions parser
# ---------------------------------------------------------------------------

def _parse_error_definitions(wb) -> dict:
    """Parse the Error Code Definition sheet into a code->name dict."""
    defs = {}
    try:
        ws = wb["Error Code Definition"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = [v for v in row if v is not None]
            if len(vals) >= 2 and isinstance(vals[1], str):
                name = str(vals[0]).strip() if vals[0] else ""
                code = vals[1].strip()
                if code and name:
                    defs[code] = name
    except KeyError:
        pass
    return defs


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def load_error_data(filepath: str | Path) -> pd.DataFrame:
    """
    Load and normalize all error data from the Fleet QC spreadsheet.
    
    Returns a DataFrame with columns:
        date, error_code, count, serial_number, unit_name,
        firmware_version, unit_type, error_category, error_name
    """
    filepath = Path(filepath)
    wb = load_workbook(filepath, read_only=True, data_only=True)

    # Parse both horizontal sheets
    all_events = []
    
    for sheet_name in wb.sheetnames:
        normalized = sheet_name.strip().lower()
        if "error code trend" in normalized:
            events = _parse_horizontal_sheet(wb[sheet_name], sheet_name)
            all_events.extend(events)

    if not all_events:
        return pd.DataFrame()

    df = pd.DataFrame(all_events)

    # Deduplicate: same date + code + count + SN is likely the same event
    # appearing in both "Error Code Trend" and "Customer Error Code Trend"
    df = df.drop_duplicates(
        subset=["date", "error_code", "count", "serial_number", "unit_name"],
        keep="first",
    )

    # Add derived columns
    df["error_category"] = df["error_code"].apply(get_error_category)
    df["error_name"] = df["error_code"].apply(get_error_name)
    df["date"] = pd.to_datetime(df["date"])
    
    # Clean up unit_type: fill gaps using name-based lookup
    # (continuation rows sometimes miss the type)
    name_to_type = df.dropna(subset=["unit_type"]).drop_duplicates("unit_name").set_index("unit_name")["unit_type"].to_dict()
    sn_to_type = df.dropna(subset=["unit_type"]).drop_duplicates("serial_number").set_index("serial_number")["unit_type"].to_dict()
    
    def fill_type(row):
        if pd.notna(row["unit_type"]):
            return row["unit_type"]
        if row["unit_name"] in name_to_type:
            return name_to_type[row["unit_name"]]
        if row["serial_number"] in sn_to_type:
            return sn_to_type[row["serial_number"]]
        return "Unknown"
    
    df["unit_type"] = df.apply(fill_type, axis=1)

    # Sort by date
    df = df.sort_values("date").reset_index(drop=True)

    # Drop the source_sheet column — it was only for dedup logic
    df = df.drop(columns=["source_sheet"], errors="ignore")

    return df


def load_firmware_updates(filepath: str | Path = None) -> pd.DataFrame:
    """
    Returns known firmware deployment dates.

    Only formally released versions are included — intermediate versions
    that were superseded without a formal deployment are excluded.
    Update this table when new firmware versions are released.

    Returns DataFrame with columns: version, first_seen_date
    """
    FIRMWARE_RELEASES = [
        ("ovvi-fw-v1.0.0", "2025-05-30"),
        ("ovvi-fw-v1.0.2", "2025-06-16"),
        ("ovvi-fw-v1.0.4", "2025-08-22"),
        ("ovvi-fw-v1.0.5", "2025-09-09"),
        ("ovvi-fw-v1.0.6", "2025-09-25"),
        ("ovvi-fw-v1.0.7", "2026-01-21"),
    ]
    df = pd.DataFrame(FIRMWARE_RELEASES, columns=["version", "first_seen_date"])
    df["first_seen_date"] = pd.to_datetime(df["first_seen_date"])
    return df


# ---------------------------------------------------------------------------
# Quick test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    
    path = sys.argv[1] if len(sys.argv) > 1 else "data/Fleet_QC_Report.xlsx"
    print(f"Loading: {path}")
    
    df = load_error_data(path)
    print(f"\nTotal events: {len(df)}")
    print(f"Date range: {df['date'].min()} to {df['date'].max()}")
    print(f"Unique units: {df['unit_name'].nunique()}")
    print(f"Unique error codes: {df['error_code'].nunique()}")
    print(f"\nBy unit type:")
    print(df.groupby("unit_type")["count"].sum())
    print(f"\nTop 10 error codes by total count:")
    print(df.groupby("error_code")["count"].sum().sort_values(ascending=False).head(10))
    print(f"\nSample rows:")
    print(df.head(10).to_string())
